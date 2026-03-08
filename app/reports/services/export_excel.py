from __future__ import annotations

from datetime import datetime
from typing import Dict

from app.utils.excel import adjust_column_widths, save_workbook_to_temp


def export_aging_report_to_excel(report_data: dict, export_dir: str) -> Dict[str, object]:
    """
    Build an Excel file for the aging report.
    Returns download metadata.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise ImportError("הספרייה openpyxl נדרשת לצורך ייצוא לאקסל. יש להתקין באמצעות: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aging Report"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"דוח חובות ללקוחות - {report_data['report_date']}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = header_alignment

    headers = [
        "שם לקוח",
        "סה\"כ חוב",
        "שוטף (0-30)",
        "30-60 ימים",
        "60-90 ימים",
        "90+ ימים",
        "תאריך עתיק",
        "ימים",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    row = 4
    for item in report_data["items"]:
        ws.cell(row=row, column=1, value=item["client_name"])
        ws.cell(row=row, column=2, value=item["total_outstanding"])
        ws.cell(row=row, column=3, value=item["current"])
        ws.cell(row=row, column=4, value=item["days_30"])
        ws.cell(row=row, column=5, value=item["days_60"])
        ws.cell(row=row, column=6, value=item["days_90_plus"])
        ws.cell(row=row, column=7, value=str(item["oldest_invoice_date"]) if item["oldest_invoice_date"] else "")
        ws.cell(row=row, column=8, value=item["oldest_invoice_days"] or "")
        row += 1

    row += 1
    summary_cell = ws.cell(row=row, column=1)
    summary_cell.value = "סיכום"
    summary_cell.font = Font(bold=True)

    ws.cell(row=row, column=2, value=report_data["total_outstanding"])
    ws.cell(row=row, column=3, value=report_data["summary"]["total_current"])
    ws.cell(row=row, column=4, value=report_data["summary"]["total_30_days"])
    ws.cell(row=row, column=5, value=report_data["summary"]["total_60_days"])
    ws.cell(row=row, column=6, value=report_data["summary"]["total_90_plus"])

    adjust_column_widths(ws)

    return save_workbook_to_temp(
        wb,
        prefix="aging_report",
        export_dir=export_dir,
        extra_meta={"format": "excel", "generated_at": datetime.now()},
    )
