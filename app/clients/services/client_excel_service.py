import tempfile
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Iterable

from sqlalchemy.orm import Session

from app.clients.constants import (
    CLIENT_EXCEL_FREEZE_PANES,
    CLIENT_EXCEL_SHEET_TITLE,
    CLIENT_EXPORT_COLUMNS,
    MAX_CLIENT_IMPORT_UPLOAD_SIZE,
    CLIENT_TEMPLATE_COLUMNS,
    CLIENT_TEMPLATE_SAMPLE_ROW,
)
from app.utils.excel import adjust_column_widths, save_workbook_to_temp

if TYPE_CHECKING:
    from app.clients.services.create_client_service import CreateClientService


class ClientExcelImportError(ValueError):
    def __init__(self, message: str, status_code: int):
        super().__init__(message)
        self.status_code = status_code


class ClientExcelService:
    """Helper for creating client Excel exports and templates."""

    def __init__(self, db: Session):
        self.db = db
        self.export_dir = Path(tempfile.gettempdir()) / "exports" / "clients"
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export_clients(self, clients: Iterable[object]) -> dict:
        """Create an Excel file containing the provided clients."""
        wb, ws = self._create_workbook_with_columns(CLIENT_EXPORT_COLUMNS)
        for row_index, client in enumerate(clients, start=2):
            for col_index, (attr, _) in enumerate(CLIENT_EXPORT_COLUMNS, start=1):
                ws.cell(row=row_index, column=col_index, value=self._value_from_client(client, attr))

        adjust_column_widths(ws)
        return save_workbook_to_temp(wb, prefix="clients_export", export_dir=self.export_dir)

    def generate_template(self) -> dict:
        """Create a client import template that includes helper headers."""
        wb, ws = self._create_workbook_with_columns(CLIENT_TEMPLATE_COLUMNS)
        for col_index, value in enumerate(CLIENT_TEMPLATE_SAMPLE_ROW, start=1):
            ws.cell(row=2, column=col_index, value=value)

        adjust_column_widths(ws)
        return save_workbook_to_temp(wb, prefix="clients_template", export_dir=self.export_dir)

    def _create_workbook_with_columns(self, columns: list[tuple[str, str]]):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
        except ImportError as exc:
            raise ImportError("הספרייה openpyxl נדרשת לצורך ייצוא לקוחות לאקסל") from exc

        wb = Workbook()
        ws = wb.active
        ws.title = CLIENT_EXCEL_SHEET_TITLE
        header_font = Font(bold=True)

        for index, (_, header) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=index, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = CLIENT_EXCEL_FREEZE_PANES
        return wb, ws

    def import_clients_from_excel(
        self,
        workbook,
        create_client_service: "CreateClientService",
        actor_id: int | None = None,
    ) -> tuple[int, list[dict]]:
        """Parse workbook and create clients with their first business."""
        worksheet = workbook.active
        created = 0
        errors: list[dict] = []

        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            values = self._template_row_values(row)
            full_name = values["full_name"]
            business_name = values["business_name"]
            id_number = values["id_number"]

            if not (full_name and business_name and id_number):
                errors.append({"row": row_index, "error": "שם מלא, שם עסק ומספר מזהה הם שדות חובה"})
                continue

            savepoint = self.db.begin_nested()
            try:
                create_client_service.create_client(
                    full_name=full_name,
                    business_name=business_name,
                    id_number=id_number,
                    phone=values["phone"] or None,
                    email=values["email"] or None,
                    actor_id=actor_id,
                )
                savepoint.commit()
                created += 1
            except Exception as exc:
                savepoint.rollback()
                errors.append({"row": row_index, "error": str(exc)})

        return created, errors

    def import_clients_from_upload(
        self,
        contents: bytes,
        create_client_service: "CreateClientService",
        actor_id: int | None = None,
        content_length: int | None = None,
        max_upload_size: int = MAX_CLIENT_IMPORT_UPLOAD_SIZE,
    ) -> dict:
        if content_length is not None and content_length > max_upload_size:
            raise ClientExcelImportError("הקובץ חורג ממגבלת הגודל של 10MB", 413)
        if len(contents) > max_upload_size:
            raise ClientExcelImportError("הקובץ חורג ממגבלת הגודל של 10MB", 413)

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ClientExcelImportError(
                "הספרייה openpyxl נדרשת לצורך ייבוא לקוחות",
                500,
            ) from exc

        try:
            workbook = load_workbook(BytesIO(contents), data_only=True)
        except Exception as exc:
            raise ClientExcelImportError("לא ניתן לקרוא את קובץ האקסל", 400) from exc

        total_rows = max(workbook.active.max_row - 1, 0)
        created, errors = self.import_clients_from_excel(
            workbook,
            create_client_service,
            actor_id=actor_id,
        )
        return {"created": created, "total_rows": total_rows, "errors": errors}

    def _value_from_client(self, client: object, attr: str):
        value = getattr(client, attr, "")
        return value or ""

    def _template_row_values(self, row) -> dict[str, str]:
        values = {}
        for index, (field_name, _) in enumerate(CLIENT_TEMPLATE_COLUMNS):
            raw_value = row[index] if len(row) > index else None
            values[field_name] = str(raw_value).strip() if raw_value is not None else ""
        return values
