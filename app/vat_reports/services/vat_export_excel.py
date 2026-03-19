"""VAT export: Excel format."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from typing import Dict

from app.utils.excel import adjust_column_widths, save_workbook_to_temp
from app.vat_reports.schemas.vat_client_summary_schema import VatPeriodRow


def export_vat_to_excel(
    client_name: str,
    business_id: int,
    year: int,
    periods: list[VatPeriodRow],
    export_dir: str,
) -> Dict[str, object]:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise ImportError("הספרייה openpyxl נדרשת. יש להתקין באמצעות: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"מע״מ {year}"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = f"{client_name} — מע״מ {year}"
    title.font = Font(bold=True, size=13)
    title.alignment = center

    headers = ["תקופה", "סטטוס", "עסקאות", "תשומות", "נטו", "סופי", "הוגש"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    row = 4
    totals = {"output": Decimal(0), "input": Decimal(0), "net": Decimal(0)}
    for p in periods:
        status_str = p.status.value if hasattr(p.status, "value") else str(p.status)
        ws.cell(row=row, column=1, value=p.period)
        ws.cell(row=row, column=2, value=status_str)
        ws.cell(row=row, column=3, value=float(p.total_output_vat))
        ws.cell(row=row, column=4, value=float(p.total_input_vat))
        ws.cell(row=row, column=5, value=float(p.net_vat))
        ws.cell(row=row, column=6, value=float(p.final_vat_amount) if p.final_vat_amount is not None else "")
        ws.cell(row=row, column=7, value=p.filed_at.strftime("%d/%m/%Y") if p.filed_at else "")
        totals["output"] += p.total_output_vat
        totals["input"] += p.total_input_vat
        totals["net"] += p.net_vat
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="סה״כ").font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(totals["output"]))
    ws.cell(row=row, column=4, value=float(totals["input"]))
    ws.cell(row=row, column=5, value=float(totals["net"]))

    adjust_column_widths(ws, max_width=40)

    return save_workbook_to_temp(
        wb,
        prefix=f"vat_{business_id}_{year}",
        export_dir=export_dir,
        extra_meta={"format": "excel", "generated_at": datetime.now()},
    )
