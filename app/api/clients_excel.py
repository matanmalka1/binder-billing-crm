import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import FileResponse

from app.api.deps import CurrentUser, DBSession, require_role
from app.models import UserRole
from app.services import ClientService
from app.services.client_excel_service import ClientExcelService

router = APIRouter(
    prefix="/clients",
    tags=["clients"],
    dependencies=[Depends(require_role(UserRole.ADVISOR, UserRole.SECRETARY))],
)

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.get("/export")
def export_clients(db: DBSession):
    """Return all clients as an Excel workbook."""
    client_service = ClientService(db)
    excel_service = ClientExcelService(db)
    try:
        result = excel_service.export_clients(client_service.list_all_clients())
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(exc),
        )

    return FileResponse(
        path=result["filepath"],
        media_type=EXCEL_MEDIA_TYPE,
        filename=result["filename"],
        headers={
            "Content-Disposition": f'attachment; filename="{result["filename"]}"',
        },
    )


@router.get("/template")
def download_client_template(db: DBSession):
    """Download a starter Excel template for client imports."""
    excel_service = ClientExcelService(db)
    try:
        result = excel_service.generate_template()
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(exc),
        )

    return FileResponse(
        path=result["filepath"],
        media_type=EXCEL_MEDIA_TYPE,
        filename=result["filename"],
        headers={
            "Content-Disposition": f'attachment; filename="{result["filename"]}"',
        },
    )


@router.post("/import")
async def import_clients_from_excel(
    file: UploadFile,
    db: DBSession,
    user: CurrentUser,
):
    """Create clients in bulk from an Excel file (advisor-only)."""
    if user.role != UserRole.ADVISOR:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only advisors can import clients",
        )

    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl is required for client import",
        ) from exc

    contents = await file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unable to read Excel file: {exc}",
        )

    worksheet = workbook.active
    total_rows = max(worksheet.max_row - 1, 0)

    client_service = ClientService(db)
    created = 0
    errors = []

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        full_name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        id_number = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        client_type = (
            str(row[2]).strip().lower() if len(row) > 2 and row[2] is not None else ""
        )
        phone = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
        email = str(row[4]).strip() if len(row) > 4 and row[4] is not None else None
        notes = str(row[5]).strip() if len(row) > 5 and row[5] is not None else None

        if not (full_name and id_number and client_type):
            errors.append(
                {
                    "row": row_index,
                    "error": "Full Name, ID Number, and Client Type are required",
                }
            )
            continue

        try:
            client_service.create_client(
                full_name=full_name,
                id_number=id_number,
                client_type=client_type,
                opened_at=date.today(),
                phone=phone,
                email=email,
                notes=notes,
            )
            created += 1
        except Exception as exc:
            errors.append({"row": row_index, "error": str(exc)})

    return {"created": created, "total_rows": total_rows, "errors": errors}
