"""
Shared openpyxl helpers used by every Excel exporter in the project.

Consumers
---------
- app/clients/services/client_excel_service.py
- app/reports/services/export_excel.py
- app/vat_reports/services/vat_export_excel.py
"""

from __future__ import annotations

import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Column widths
# ---------------------------------------------------------------------------


def adjust_column_widths(
    ws,
    *,
    min_width: int = 8,
    max_width: int = 50,
    padding: int = 2,
) -> None:
    """
    Auto-fit every column's width to its widest cell value.

    Parameters
    ----------
    ws        : openpyxl Worksheet
    min_width : floor so narrow columns (e.g. booleans) are still readable.
    max_width : ceiling so prose columns don't stretch the sheet.
    padding   : extra characters added on top of the measured max length.
    """
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:  # openpyxl not installed — skip silently
        return

    for idx, col_cells in enumerate(ws.iter_cols(), start=1):
        measured = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0,
        )
        width = max(min_width, min(measured + padding, max_width))
        ws.column_dimensions[get_column_letter(idx)].width = width


# ---------------------------------------------------------------------------
# Save workbook
# ---------------------------------------------------------------------------


def save_workbook_to_temp(
    wb,
    prefix: str,
    *,
    subdir: str = "exports",
    export_dir: str | Path | None = None,
    extra_meta: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """
    Save *wb* to a temp directory and return download metadata.

    Parameters
    ----------
    wb          : openpyxl Workbook
    prefix      : filename prefix, e.g. ``"clients_export"``
    subdir      : sub-folder inside the system temp dir (used when
                  *export_dir* is not supplied). Default: ``"exports"``.
    export_dir  : explicit target directory; overrides *subdir* when given.
                  Accepts ``str`` or ``pathlib.Path``.
    extra_meta  : additional key/value pairs merged into the returned dict.
                  Callers that need ``"format"`` or ``"generated_at"`` can
                  pass them here, e.g.::

                      extra_meta={"format": "excel", "generated_at": datetime.now()}

    Returns
    -------
    dict with at minimum ``{"filepath": str, "filename": str}``, plus
    whatever was supplied in *extra_meta*.
    """
    if export_dir is not None:
        target_dir = Path(export_dir)
    else:
        target_dir = Path(tempfile.gettempdir()) / subdir

    target_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{timestamp}.xlsx"
    filepath = target_dir / filename
    wb.save(filepath)

    result: dict[str, Any] = {"filepath": str(filepath), "filename": filename}
    if extra_meta:
        result.update(extra_meta)
    return result
