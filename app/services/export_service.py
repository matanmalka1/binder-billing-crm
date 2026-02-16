import io
import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from sqlalchemy.orm import Session


class ExportService:
    """
    Report export service for Excel and PDF.
    
    Generates downloadable files from report data.
    """

    def __init__(self, db: Session):
        self.db = db
        export_base = Path(tempfile.gettempdir()) / "exports"
        export_base.mkdir(parents=True, exist_ok=True)
        self.export_dir = str(export_base)

    def export_aging_report_to_excel(self, report_data: dict) -> dict:
        """
        Export aging report to Excel format.
        
        Returns download info.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aging Report"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Title
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"דוח חובות ללקוחות - {report_data['report_date']}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = header_alignment

        # Column headers
        headers = ["שם לקוח", "סה\"כ חוב", "שוטף (0-30)", "30-60 ימים", "60-90 ימים", "90+ ימים", "תאריך עתיק", "ימים"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Data rows
        row = 4
        for item in report_data["items"]:
            ws.cell(row=row, column=1, value=item["client_name"])
            ws.cell(row=row, column=2, value=item["total_outstanding"])
            ws.cell(row=row, column=3, value=item["current"])
            ws.cell(row=row, column=4, value=item["days_30"])
            ws.cell(row=row, column=5, value=item["days_60"])
            ws.cell(row=row, column=6, value=item["days_90_plus"])
            ws.cell(row=row, column=7, value=str(item["oldest_invoice_date"]) if item["oldest_invoice_date"] else "")
            ws.cell(row=row, column=8, value=item["oldest_invoice_days"] or "")
            row += 1

        # Summary row
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        summary_cell = ws.cell(row=row, column=1)
        summary_cell.value = "סיכום"
        summary_cell.font = Font(bold=True)
        
        ws.cell(row=row, column=2, value=report_data["total_outstanding"])
        ws.cell(row=row, column=3, value=report_data["summary"]["total_current"])
        ws.cell(row=row, column=4, value=report_data["summary"]["total_30_days"])
        ws.cell(row=row, column=5, value=report_data["summary"]["total_60_days"])
        ws.cell(row=row, column=6, value=report_data["summary"]["total_90_plus"])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to file
        filename = f"aging_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        wb.save(filepath)

        return {
            "download_url": f"/exports/{filename}",
            "filename": filename,
            "format": "excel",
            "generated_at": datetime.now(),
        }

    def export_aging_report_to_pdf(self, report_data: dict) -> dict:
        """
        Export aging report to PDF format.
        
        Returns download info.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")

        # Create PDF
        filename = f"aging_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.export_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph(f"<b>דוח חובות ללקוחות - {report_data['report_date']}</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.5*cm))

        # Table data
        table_data = [
            ["שם לקוח", "סה\"כ חוב", "שוטף (0-30)", "30-60 ימים", "60-90 ימים", "90+ ימים", "תאריך עתיק", "ימים"]
        ]

        for item in report_data["items"]:
            table_data.append([
                item["client_name"],
                f"{item['total_outstanding']:.2f}",
                f"{item['current']:.2f}",
                f"{item['days_30']:.2f}",
                f"{item['days_60']:.2f}",
                f"{item['days_90_plus']:.2f}",
                str(item["oldest_invoice_date"]) if item["oldest_invoice_date"] else "",
                str(item["oldest_invoice_days"]) if item["oldest_invoice_days"] else "",
            ])

        # Summary row
        table_data.append([
            "סיכום",
            f"{report_data['total_outstanding']:.2f}",
            f"{report_data['summary']['total_current']:.2f}",
            f"{report_data['summary']['total_30_days']:.2f}",
            f"{report_data['summary']['total_60_days']:.2f}",
            f"{report_data['summary']['total_90_plus']:.2f}",
            "",
            "",
        ])

        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)

        return {
            "download_url": f"/exports/{filename}",
            "filename": filename,
            "format": "pdf",
            "generated_at": datetime.now(),
        }
