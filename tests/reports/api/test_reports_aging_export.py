import io
from datetime import date, timedelta
from decimal import Decimal

import openpyxl

from app.charge.models.charge import Charge, ChargeStatus, ChargeType
from tests.helpers.identity import seed_business, seed_client_identity


def _seed_charges(db):
    client = seed_client_identity(
        db,
        full_name="Export Aging Client",
        id_number="AGING-EXP-1",
    )
    business = seed_business(
        db,
        legal_entity_id=client.legal_entity_id,
        business_name=client.full_name,
        opened_at=date.today(),
    )
    db.commit()
    db.refresh(business)

    issued_at = date.today() - timedelta(days=40)
    charge = Charge(
        client_record_id=client.id,
        business_id=business.id,
        amount=Decimal("250.00"),
        charge_type=ChargeType.CONSULTATION_FEE,
        status=ChargeStatus.ISSUED,
        issued_at=issued_at,
        created_at=issued_at,
    )
    db.add(charge)
    db.commit()
    db.refresh(charge)
    return client


def test_aging_report_export_excel(client, test_db, advisor_headers):
    crm_client = _seed_charges(test_db)

    resp = client.get("/api/v1/reports/aging/export?format=excel", headers=advisor_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # Title row merged A1:H1
    assert ws["A1"].value is not None
    # Data row for client
    client_names = [ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)]
    assert crm_client.full_name in client_names


def test_aging_report_export_pdf(client, test_db, advisor_headers):
    _seed_charges(test_db)

    resp = client.get("/api/v1/reports/aging/export?format=pdf", headers=advisor_headers)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/pdf")
    assert len(resp.content) > 0
