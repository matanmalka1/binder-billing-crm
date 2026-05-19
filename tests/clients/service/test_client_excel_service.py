from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.clients.models.client_record import ClientRecord
from app.clients.services.client_excel_service import ClientExcelService
from app.clients.services.client_query_service import ClientQueryService
from app.clients.services.create_client_service import create_client_identity_only
from app.common.enums import AdvancePaymentFrequency, EntityType, IdNumberType, VatType


def test_client_excel_import_skips_blank_and_collects_errors(test_db):
    service = ClientExcelService(test_db)
    wb = Workbook()
    ws = wb.active
    ws.append(["full_name", "business_name", "id_number", "phone", "email"])
    ws.append([None, None, None, None, None])
    ws.append(["", "", "", "", ""])
    ws.append(["Good Name", "Good Business", "ID1", "", ""])

    class _ClientSvc:
        def create_client(self, **kwargs):
            if kwargs["id_number"] == "ID1":
                raise RuntimeError("duplicate")

    created, errors = service.import_clients_from_excel(wb, _ClientSvc(), actor_id=1)

    assert created == 0
    assert len(errors) == 1
    assert errors[0]["row"] == 4


def test_client_excel_import_collects_required_field_errors(test_db):
    service = ClientExcelService(test_db)
    wb = Workbook()
    ws = wb.active
    ws.append(["full_name", "business_name", "id_number", "phone", "email"])
    ws.append(["Missing ID", "Missing ID Business", "", "", ""])
    ws.append(["Missing Business", "", "760000001", "", ""])
    ws.append(["", "Missing Name Business", "760000002", "", ""])

    class _ClientSvc:
        def create_client(self, **_kwargs):
            raise AssertionError("create should not be called for invalid rows")

    created, errors = service.import_clients_from_excel(wb, _ClientSvc(), actor_id=1)

    assert created == 0
    assert len(errors) == 3
    assert {err["row"] for err in errors} == {2, 3, 4}


def test_client_excel_import_uses_tax_defaults_when_optional_columns_are_blank(test_db):
    service = ClientExcelService(test_db)
    wb = Workbook()
    ws = wb.active
    ws.append(["full_name", "business_name", "id_number", "phone", "email"])
    ws.append(["Imported Name", "Imported Business", "123456789", "", ""])
    calls = []

    class _ClientSvc:
        def create_client(self, **kwargs):
            calls.append(kwargs)

    created, errors = service.import_clients_from_excel(wb, _ClientSvc(), actor_id=1)

    assert created == 1
    assert errors == []
    assert calls[0]["entity_type"] == EntityType.OSEK_MURSHE
    assert calls[0]["vat_reporting_frequency"] == VatType.BIMONTHLY
    assert calls[0]["advance_payment_frequency"] == AdvancePaymentFrequency.BIMONTHLY


def test_client_excel_import_rolls_back_failed_create_client_row(test_db):
    service = ClientExcelService(test_db)
    wb = Workbook()
    ws = wb.active
    ws.append(["full_name", "business_name", "id_number", "phone", "email"])
    ws.append(["Half Created", "Half Created Business", "ROLLBACK-1", "", ""])

    class _ClientSvc:
        def create_client(self, **kwargs):
            create_client_identity_only(
                test_db,
                full_name=kwargs["full_name"],
                id_number=kwargs["id_number"],
                id_number_type=IdNumberType.INDIVIDUAL,
            )
            raise RuntimeError("business failed")

    created, errors = service.import_clients_from_excel(wb, _ClientSvc(), actor_id=1)

    assert created == 0
    assert len(errors) == 1
    assert test_db.scalar(select(func.count(ClientRecord.id))) == 0


def test_client_excel_export_and_template_generate_files(test_db):
    service = ClientExcelService(test_db)

    client_record = create_client_identity_only(
        test_db,
        full_name="Excel Name",
        id_number="750000001",
        id_number_type=IdNumberType.CORPORATION,
        entity_type=EntityType.COMPANY_LTD,
        phone="0501234567",
        email="excel@test.com",
        actor_id=1,
    )
    test_db.commit()
    full_client = ClientQueryService(test_db).get_full_client(client_record.id)

    exported = service.export_clients([full_client])
    template = service.generate_template()

    assert Path(exported["filepath"]).exists()
    assert exported["filename"].endswith(".xlsx")

    wb = load_workbook(exported["filepath"])
    ws = wb.active
    assert ws.cell(row=1, column=2).value == "Full Name"
    assert ws.cell(row=2, column=2).value == "Excel Name"

    assert Path(template["filepath"]).exists()
    template_wb = load_workbook(template["filepath"])
    template_ws = template_wb.active
    assert template_ws.cell(row=1, column=1).value == "Full Name"
    assert template_ws.cell(row=2, column=1).value == "יוסי כהן"
    assert template_ws.cell(row=1, column=2).value == "Business Name"
    assert template_ws.cell(row=2, column=2).value == "יוסי כהן ייעוץ"
    assert template_ws.cell(row=1, column=6).value == "Entity Type (optional)"
    assert template_ws.cell(row=2, column=6).value == "osek_murshe"

