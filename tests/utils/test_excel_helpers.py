import tempfile
from pathlib import Path

from openpyxl import Workbook

from app.utils.excel import adjust_column_widths, save_workbook_to_temp


def test_adjust_column_widths_apply_styles():
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "Long Header"])
    ws.append(["short", "very long value here"])

    adjust_column_widths(ws, min_width=5, max_width=40, padding=1)

    assert ws.column_dimensions["B"].width >= ws.column_dimensions["A"].width


def test_save_workbook_to_temp_returns_metadata(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["x"])

    payload = save_workbook_to_temp(
        wb,
        prefix="unit_excel",
        export_dir=tmp_path,
        extra_meta={"format": "excel"},
    )

    assert payload["format"] == "excel"
    assert payload["filename"].startswith("unit_excel_")
    assert Path(payload["filepath"]).exists()


def test_save_workbook_to_temp_uses_default_subdir():
    wb = Workbook()
    ws = wb.active
    ws.append(["x"])
    payload = save_workbook_to_temp(wb, prefix="default_subdir_case")
    assert payload["filepath"].startswith(str(Path(tempfile.gettempdir()) / "exports"))
